from py2neo import Graph, Node, Path, Relationship
import openpyxl
# References on https://py2neo.org/2.0/essentials.html

class ConectarNeo4J():
    __graph = Graph()

    def __init__(self):
        self.__graph = Graph("bolt://35.237.186.164:7687", bolt=True, auth=("neo4j", "stagiopbd"))

    def limpar(self):
        tx = self.__graph.begin()
        tx.run("MATCH (n) DETACH DELETE n")
        tx.commit()

    def returnDataset(self, cypher):
        tx = self.__graph.begin()
        for record in tx.run(cypher):
            print(record["n"])

    def Area(self, id, name):
        tx = self.__graph.begin()

        area = Node("Area")
        area["id"] = id
        area["name"] = name
        tx.create(area)
        tx.commit()

    def Zipcode(self, id, code, street):
        tx = self.__graph.begin()

        zipcode = Node("Zipcode")
        zipcode["id"] = id
        zipcode["code"] = code
        zipcode["street"] = street
        tx.create(zipcode)
        tx.commit()

    def Address(self, id, number, complement):
        tx = self.__graph.begin()

        address = Node("Address")
        address["id"] = id
        address["name"] = str(number) + " " + str(complement)
        address["number"] = number
        address["complement"] = complement
        tx.create(address)
        tx.commit()

    def Patient(self, cpf, name, birthdate, gender, bloodtype, email):
        tx = self.__graph.begin()

        patient = Node("Patient")
        patient["cpf"] = str(cpf)
        patient["name"] = name
        patient["birthdate"] = birthdate
        patient["gender"] = gender
        patient["bloodtype"] = bloodtype
        patient["email"] = email
        tx.create(patient)
        tx.commit()

    def HasArea(self, zipcode, area):
        tx = self.__graph.begin()
        tx.run("MATCH (z:Zipcode), (a:Area) WHERE z.code = '" + str(zipcode) + "' and a.id = " + str(area) + " CREATE (z)-[h:HAS_AREA]->(a)")
        tx.commit()

    def HasZipcode(self, address, zipcode):
        tx = self.__graph.begin()
        tx.run("MATCH (a:Address), (z:Zipcode) WHERE a.id = " + str(address) + " and z.code = '" + str(zipcode) + "'  CREATE (a)-[h:HAS_ZIPCODE]->(z)")
        tx.commit()

    def HasAddress(self, patient, address):
        tx = self.__graph.begin()
        tx.run("MATCH (p:Patient), (a:Address) WHERE p.cpf = '" + str(patient) + "' and a.id = " + str(address) + " CREATE (p)-[h:HAS_ADDRESS]->(a)")
        tx.commit()

if __name__ == "__main__":
    conectar = ConectarNeo4J()

    conectar.limpar()

    workbook = openpyxl.load_workbook("NEO4J-patient.xlsx", True)

    sheet = workbook['area']
    for i in range(2, sheet.max_row + 1):
        id = sheet["B" + str(i)].value
        name = sheet["A" + str(i)].value
        conectar.Area(id, name)

    sheet = workbook['zipcode']
    for i in range(2, sheet.max_row + 1):
        id = sheet["A" + str(i)].value
        code = sheet["B" + str(i)].value
        street = sheet["C" + str(i)].value
        area = sheet["D" + str(i)].value
        conectar.Zipcode(id, code, street)
        conectar.HasArea(code, area)

    sheet = workbook['address']
    for i in range(2, sheet.max_row + 1):
        id = sheet["A"+str(i)].value
        number = sheet["B"+str(i)].value
        complement = sheet["C"+str(i)].value
        zipcode = sheet["D"+str(i)].value
        conectar.Address(id, number, complement)
        conectar.HasZipcode(id, zipcode)

    sheet = workbook['patient']
    for i in range(2, sheet.max_row + 1):
        cpf = sheet["A"+str(i)].value
        name = sheet["B"+str(i)].value
        birthdate = sheet["C"+str(i)].value
        gender = sheet["D"+str(i)].value
        bloodtype = sheet["E"+str(i)].value
        email = sheet["F"+str(i)].value
        address = sheet["G"+str(i)].value
        conectar.Patient(cpf, name, birthdate, gender, bloodtype, email)
        conectar.HasAddress(cpf, address)